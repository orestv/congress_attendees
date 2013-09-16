# -*- coding: utf-8 -*
import os
from openpyxl import Workbook
from tempfile import mkdtemp
import fields

def cleanup(filename):
	pass

def export(attendees, events, users):
	wb = Workbook()
	export_attendees(wb, attendees, events, users)
	tmpd = mkdtemp()
	filename = 'report.xlsx'
	path = os.path.join(tmpd, filename)
	print path
	wb.save(path)
	return path


def export_attendees(workbook, attendees, events, users):
	ws = workbook.get_active_sheet()
	header_items = [field['caption'] for field in fields.INFO_FIELDS] + \
		[event['caption'] for event in events] + [u'Зареєстрував']
	header_row(ws, header_items)
	row = 1
	for attendee in attendees:
		column = 0
		items = [attendee.get(field['fieldId'], '') for field in fields.INFO_FIELDS]
		items = map(format_field, items)
		items += map(lambda evt: event_status(attendee['attended_events'], str(evt['_id'])), 
			events)
		uid = attendee.get('registered_by', None)
		if uid:
			user = [u['firstname'] + u['lastname'] for u in users
				if str(u['_id']) == uid][0]
			items += user
		data_row(ws, row, items)
		row += 1
	ws.title = u'Учасники'

def event_status(attended_events, eid):
	a_evts = filter(lambda e : e['_id'] == eid,
		attended_events)
	if a_evts:
		a_evt = a_evts[0]
		if a_evt.get('paid', False):
			return u'Оплачено'
		else:
			return u'Заброньовано'

def format_field(field):
	if isinstance(field, bool):
		return u'Так' if field else u'Ні'
	return field

def data_row(ws, row, items):
	column = 0
	for item in items:
		cell = ws.cell(row=row, column=column)
		cell.value = item
		column += 1

def header_row(ws, items):
	row = 0
	column = 0
	for item in items:
		cell = ws.cell(row=row, column=column)
		cell.style.font.bold = True
		cell.value = item
		column += 1